"""Handles excel sample sheet import and export as well as proposal ID validation. 

Contains code to parse excel sheet for bar (list of sample dict), export parsed excel sheet, and export excel metadata rules for mediaWiki.
"""

# imports
from copy import deepcopy
from openpyxl import load_workbook
from openpyxl.writer import excel
from pathlib import Path
from datetime import date, datetime
import json
import re, warnings, httpx, uuid
import numpy as np
import pandas as pd
from .defaults import (
    rsoxs_configurations,
    empty_sample,
    empty_acq,
    edge_names,
    config_list,
    current_version,
    rsoxs_edges,
    nexafs_edges,
)


def load_samplesxlsx(filename: str, verbose=False):
    """Imports data from sample excel spreadsheet and online sources to generate bar (list of sample dicts)

    Parameters
    ----------
    filename : str
        String or Pathlike object that references the excel sheet to load
    verbose : bool
        Whether or not to print messages to stdout

    Returns
    -------
    list of dicts
        bar (list of sample dicts) which contain all imported data from the Bar sheet and Acquisitions sheet
    """

    # Validate Spreadsheet Version Number
    excel_file = load_workbook(filename)
    print(f"spreadsheet version is {excel_file.properties.title}")
    if excel_file.properties.title != current_version:
        excel_file.close()
        raise ValueError(
            "this excel file is not the current version, we read{.  please upgrade your template"
            " and try again"
        )
    excel_file.close()

    # First, check the bar sheet for whether header rows with user instructions are present, and identity them if so
    # If so, we can do some extra validation, but need to skip them when loading data
    # If not, we proceed with a bit less validation and don't skip them
    barHeaderRows = []
    barParamsRequired = (
        []
    )  # if reading the header fails, this stays empty and we don't check for them

    try:
        # Import just where the header rows from the bar sheet would be
        warnings.simplefilter(action="ignore", category=UserWarning)
        if verbose:
            print("Loading 'Bar' Sheet Headers")
        df_barHeader = pd.read_excel(
            filename,
            na_values="",
            engine="openpyxl",
            keep_default_na=True,
            converters={"sample_date": str},
            sheet_name="Bar",
            nrows=4,
            verbose=verbose,
        )

        # Check if the first cell has the Parameters'Index
        if "Parameter/ Index" in df_barHeader.keys():
            doImportBarHeaders = True
        else:  # Dont attempt to import header and warn user that we will skip some cell validation
            doImportBarHeaders = False
            # warnings.resetwarnings()
            # warnings.warn(
            #    "\nDidn't find Parameter/ Index column in bar sheet, skipping some validation that needs header cells"
            # )

        # Load the header data
        if doImportBarHeaders:
            # Check where we actually have the header rows. Mark them to be skipped when loading data.
            if df_barHeader["Parameter/ Index"][0] != "Description":
                raise ValueError("Couldn't find parameter Descriptions")
            else:
                barHeaderRows.append(1)
            if (
                len(df_barHeader["Parameter/ Index"]) > 1
                and df_barHeader["Parameter/ Index"][1] == "Rules"
            ):
                barHeaderRows.append(2)
            if (
                len(df_barHeader["Parameter/ Index"]) > 2
                and df_barHeader["Parameter/ Index"][2] == "Example"
            ):
                barHeaderRows.append(3)
            if (
                len(df_barHeader["Parameter/ Index"]) > 3
                and df_barHeader["Parameter/ Index"][3] == "Notes"
            ):
                barHeaderRows.append(4)

            # Cast as a nested dict, where:
            dict_barHeader = df_barHeader.to_dict(orient="index")

            # Make a list of bar parameters that are marked 'REQUIRED'
            barParamsRequired = []
            for param in dict_barHeader[0]:
                if "REQUIRED" in dict_barHeader[0][param]:
                    barParamsRequired.append(param)

    except ValueError as e:
        warnings.resetwarnings()
        warnings.warn(
            (
                "\nError parsing bar sheet headers, skipping some validation that needs header"
                f" cells: {str(e)}"
            ),
            stacklevel=2,
        )
        pass

    # Then, check the Acquisitions sheet for whether header rows with user instructions are present, and identity them if so
    # If so, we can do some extra validation, but need to skip them when loading data
    # If not, we proceed with a bit less validation and don't skip them
    acqHeaderRows = []
    acqParamsRequired = (
        []
    )  # if reading the header fails, this stays empty and we don't check for them

    try:
        # Import just where the header rows from the Acquisitions sheet would be
        warnings.simplefilter(action="ignore", category=UserWarning)
        if verbose:
            print("Loading 'Acquisitions' Sheet Headers")
        df_acqHeader = pd.read_excel(
            filename,
            na_values="",
            engine="openpyxl",
            keep_default_na=True,
            converters={"sample_date": str},
            sheet_name="Acquisitions",
            nrows=4,
            verbose=verbose,
        )

        # Check if the first cell has the Parameters'Index
        if "Parameter/ Index" in df_acqHeader.keys():
            doImportAcqHeaders = True
        else:  # Dont attempt to import header and warn user that we will skip some cell validation
            doImportAcqHeaders = False
            # warnings.resetwarnings()
            # warnings.warn(
            #    "\nDidn't find Parameter/ Index column in acq sheet, skipping some validation that needs header cells"
            # )

        # Load the header data
        if doImportBarHeaders:
            # Check where we actually have the header rows. Mark them to be skipped when loading data.
            if df_acqHeader["Parameter/ Index"][0] != "Description":
                raise ValueError("Couldn't find parameter Descriptions")
            else:
                acqHeaderRows.append(1)
            if (
                len(df_acqHeader["Parameter/ Index"]) > 1
                and df_acqHeader["Parameter/ Index"][1] == "Rules"
            ):
                acqHeaderRows.append(2)
            if (
                len(df_acqHeader["Parameter/ Index"]) > 2
                and df_acqHeader["Parameter/ Index"][2] == "Example"
            ):
                acqHeaderRows.append(3)
            if (
                len(df_acqHeader["Parameter/ Index"]) > 3
                and df_acqHeader["Parameter/ Index"][3] == "Notes"
            ):
                acqHeaderRows.append(4)

            # Cast as a nested dict, where:
            dict_acqHeader = df_acqHeader.to_dict(orient="index")

            # Make a list of acq parameters that are marked 'REQUIRED'
            acqParamsRequired = []
            for param in dict_acqHeader[0]:
                if "REQUIRED" in dict_acqHeader[0][param]:
                    acqParamsRequired.append(param)

    except ValueError as e:
        warnings.resetwarnings()
        warnings.warn(
            (
                "\nError parsing Acquisitions sheet headers, skipping some validation that needs"
                f" header cells: {str(e)}"
            ),
            stacklevel=2,
        )
        pass

    # Import Bar sheet data cells as a dataframe
    if verbose:
        print("Loading 'Bar' Sheet Data")
    warnings.simplefilter(action="ignore", category=UserWarning)
    df_bar = pd.read_excel(
        filename,
        na_values="",
        engine="openpyxl",
        keep_default_na=True,
        converters={"sample_date": str},
        sheet_name="Bar",
        skiprows=barHeaderRows,
        verbose=verbose,
    )

    # Replace NaNs with empty string
    df_bar.replace(np.nan, "", regex=True, inplace=True)

    # Convert dataframe to a list of dictionaries, each row is a list element and each column is a key->value pair
    new_bar = df_bar.to_dict(orient="records")

    # if the bar has one element, force to a list
    if not isinstance(new_bar, list):
        new_bar = [new_bar]

    # blank out any acquisitions elements which might be there (they shouldn't be there unless someone added a column for some reason
    for samp in new_bar:
        samp["acquisitions"] = []

    # Import Acquisitions sheet data cells as a dataframe
    if verbose:
        print("Loading 'Acquisitions' Sheet Data")
    df_acqs = pd.read_excel(
        filename,
        na_values="",
        engine="openpyxl",
        keep_default_na=True,
        sheet_name="Acquisitions",
        skiprows=acqHeaderRows,
        verbose=True,
    )

    # acqsdf.replace(np.nan, "", regex=True, inplace=True)

    # Convert dataframe to a list of dictionaries, each row is a list element and each column is a key->value pair
    acqs = df_acqs.to_dict(orient="records")

    # if the acqs has one element, force to a list
    if not isinstance(acqs, list):
        acqs = [acqs]

    if verbose:
        print("Started Parsing Acquisition Data")
    acq = {}
    # Loop through acquisitions and sanitize / validate user input
    for i, acq in enumerate(acqs):
        # Loop through columns in the acquisition sheet and sanitize strings
        for key in acq:
            # If cell holds a string, for to square brackets and single quotes
            if isinstance(acq[key], str):
                acq[key] = acq[key].replace("(", "[").replace(")", "]").replace("'", '"')
                # Try to parse values
                try:
                    acq[key] = json.loads(acq[key])
                except:
                    pass  ### TODO handle this?
            if isinstance(acq[key], str):
                if "," in acq[key]:  # if the string looks like a list
                    try:
                        acq[key] = [
                            float(num) for num in acq[key].split(",")
                        ]  # cast it as a list of floating point numbers instead
                    except:
                        pass

        # Check if values were provided for all 'REQUIRED' acquisition cells for this acq
        missedVal = False
        missingValText = (
            f"Acquisition #{i}, sample_id:{acq['sample_id']} is missing REQUIRED Parameters: "
        )
        for key in acq:  # Loop through all columns for this acquisition
            # first check if required cell, then whether it is empty
            # Empty cells are cells are stored as np.nan. isnan cant evaluate on all datatypes though
            if key in acqParamsRequired and isinstance(acq[key], float) and np.isnan(acq[key]):
                missingValText += f"{key}, "
                missedVal = True
        if missedVal:
            raise ValueError(missingValText)

        # get the sample that corresponds to the sample_id for this acq... the first one that matches it takes
        try:
            samp = next(dict for dict in new_bar if dict["sample_id"] == acq["sample_id"])
        except StopIteration:
            pass
            missingSampText = (
                f'ERROR acquisition #{i} needs a sample_id "{acq["sample_id"]}" which was not'
                " found - please check your sample_ids"
            )
            raise ValueError(missingSampText)

        acq = {key: val for key, val in acq.items() if val == val and val != ""}

        # Parse edge
        try:
            acq["edge"] = json.loads(acq["edge"])
        except:  # if edge isn't json parsable, it's probably just a string, and that's fine
            pass
        if isinstance(acq["edge"], str):
            if "," in acq["edge"]:  # if the string looks like a list
                acq["edge"] = [
                    float(num) for num in acq["edge"].split(",")
                ]  # cast it as a list of floating point numbers instead

        # Validate based on scan type
        # Validate RSoXS
        if acq["type"].lower() == "rsoxs":
            # Validate rsoxs configuration
            if acq["configuration"] not in rsoxs_configurations:
                raise TypeError(
                    f'{acq["configuration"]} on line {i} is not a valid configuration for an rsoxs'
                    " scan"
                )
            # Validate rsoxs edge
            if not isinstance(acq.get("edge", "c"), (tuple, list, int, float)):
                if not str(acq.get("edge", "c")).lower() in nexafs_edges.keys():
                    if not edge_names[str(acq.get("edge", "c")).lower()] in rsoxs_edges.keys():
                        raise ValueError(
                            f'{acq["edge"]} on line {i} is not a valid edge for an rsoxs scan'
                        )
        # Validate NEXAFS
        elif acq["type"].lower() == "nexafs":
            # Validate nexafs configuration
            if acq["configuration"] not in config_list:
                raise TypeError(
                    f'{acq["configuration"]} on line {i} is not a valid configuration for a nexafs'
                    " scan"
                )
            # Validate nexafs edge
            if not isinstance(acq.get("edge", "c"), (tuple, list)):
                if not str(acq.get("edge", "c")).lower() in nexafs_edges.keys():
                    if not edge_names[str(acq.get("edge", "c")).lower()] in nexafs_edges.keys():
                        raise ValueError(
                            f'{acq["edge"]} on line {i} is not a valid edge for a nexafs scan'
                        )

        # Validate step NEXAFS
        elif acq["type"].lower() == "nexafs":
            # Validate nexafs configuration
            if acq["configuration"] not in config_list:
                raise TypeError(
                    f'{acq["configuration"]} on line {i} is not a valid configuration for a nexafs'
                    " step scan"
                )
            # Validate nexafs edge
            if not isinstance(acq.get("edge", "c"), (tuple, list)):
                if not str(acq.get("edge", "c")).lower() in edge_names.keys():
                    raise ValueError(
                        f'{acq["edge"]} on line {i} is not a valid edge for a nexafs scan'
                    )

        # Encapsulate single-element values into lists
        if "polarizations" in acq:
            if isinstance(acq["polarizations"], (int, float)):
                acq["polarizations"] = [acq["polarizations"]]
        if "angles" in acq:
            if isinstance(acq["angles"], (int, float)):
                acq["angles"] = [acq["angles"]]
        if "temperatures" in acq:
            if isinstance(acq["temperatures"], (int, float)):
                acq["temperatures"] = [acq["temperatures"]]

        # Sanitize Group
        if not isinstance(acq.get("group", 0), str):
            acq["group"] = str(acq.get("group", ""))
        acq["uid"] = str(uuid.uuid1())
        samp["acquisitions"].append(acq)

        # sanatize grating
        if "grating" in acq:
            if isinstance(acq["grating"], (float)):
                acq["grating"] = int(acq["grating"])

        # Validate Acquisition Parameters by Value and issue warnings if out of bounds
        invalidAcqParam = False  # False means none invalid
        invalidAcqParamText = (
            f"Acquisition #{i}, sample_id:{acq['sample_id']} has invalid parameters: \n"
        )

        # Check each angle listed, if any
        if "angles" in acq:
            for testAng in acq["angles"]:
                # print(f"grazing: {samp['grazing']}, testAng: {testAng}"
                angInvalid = False
                # No need to test if its empty (nan)
                if isinstance(testAng, float) and np.isnan(testAng):
                    pass
                else:
                    # if grazing is true, valid angles are 20 to 90 degrees
                    if samp["grazing"] == True:
                        gAngLow, gAngHigh = 20, 90  # Eventually these should move to defaults page
                        try:
                            angInvalid = not isParamValid(
                                testAng, matchType="numeric", validValues=[gAngLow, gAngHigh]
                            )
                        except ValueError:  # hacky check for param type validation
                            angInvalid = True
                        if angInvalid:
                            invalidAcqParamText += (
                                f"\tInvalid Angle: {testAng} . For grazing samples, angle must be"
                                f" between {gAngLow} and {gAngHigh}\n"
                            )

                    elif samp["grazing"] == False:
                        tAngLow, tAngHigh = (
                            -14,
                            90,
                        )  # Eventually these should move to defaults page
                        try:
                            angInvalid = not isParamValid(
                                testAng, matchType="numeric", validValues=[tAngLow, tAngHigh]
                            )
                        except ValueError:  # hacky check for param type validation
                            angInvalid = True
                        if angInvalid:
                            invalidAcqParamText += (
                                f"\tInvalid Angle: {testAng} . For transmission samples, angle"
                                f" must be between {tAngLow} and {tAngHigh}\n"
                            )
                    else:
                        angInvalid = True
                        raise ValueError("Unable to determine angle geometry, check grazing field")

                # if any invalid, need to report out
                invalidAcqParam = (
                    invalidAcqParam or angInvalid
                )  # we already had invalids, or we found one now

        if invalidAcqParam:
            warnings.resetwarnings()
            warnings.warn(invalidAcqParamText, stacklevel=2)

    # Begin Bar validation
    if verbose:
        print("Started Parsing Bar Data")

    # Loop through samples in Bar and sanitize / validate user input
    for i, sam in enumerate(new_bar):
        # Handle the autogenerated columns,
        if sam.get("location", "") == "":
            new_bar[i]["location"] = "[]"
        new_bar[i]["location"] = json.loads(sam.get("location", "[]").replace("'", '"'))
        if sam.get("bar_loc", "") == "":
            new_bar[i]["bar_loc"] = "{}"
        new_bar[i]["bar_loc"] = json.loads(sam.get("bar_loc", "{}").replace("'", '"'))
        if sam.get("acq_history", "") == "":
            new_bar[i]["acq_history"] = "[]"
        new_bar[i]["acq_history"] = json.loads(
            sam.get("acq_history", "[]").replace("'", '"').rstrip('\\"').lstrip('\\"')
        )

        # Check if values were provided for all 'REQUIRED' bar cells for this sample_id
        missedVal = False
        missingValText = (
            f"Bar Entry #{i}, sample_id: {sam['sample_id']} is missing REQUIRED Parameters: "
        )
        for key in sam:  # Loop through all columns for this acquisition
            # first check if required cell, then whether it is empty
            # Empty cells are stored as empty strings
            if key in barParamsRequired and sam[key] == "":
                missingValText += f"{key}, "
                missedVal = True
        if missedVal:
            raise ValueError(missingValText)

        # Validate Bar Parameters by Value and issue warnings if out of bounds
        invalidAcqParam = False  # False means none invalid
        invalidAcqParamText = (
            f"Bar Sheet Entry #{i}, sample_id:{sam['sample_id']} has invalid parameters: \n"
        )

        # Check angle listed, if any
        if sam["angle"] != "":
            testAng = sam["angle"]
            # print(f"grazing: {sam['grazing']}, testAng: {testAng}")
            angInvalid = False
            # No need to test if its empty (nan)
            if isinstance(testAng, float) and np.isnan(testAng):
                pass
            else:
                # if grazing is true, valid angles are 20 to 90 degrees
                if sam["grazing"] == True:
                    gAngLow, gAngHigh = 20, 90  # Eventually these should move to defaults page
                    try:
                        angInvalid = not isParamValid(
                            testAng, matchType="numeric", validValues=[gAngLow, gAngHigh]
                        )
                    except ValueError:  # hacky check for param type validation
                        angInvalid = True
                    if angInvalid:
                        invalidAcqParamText += (
                            f"\tInvalid Angle: {testAng} . For grazing samples, angle must be"
                            f" between {gAngLow} and {gAngHigh}\n"
                        )

                elif sam["grazing"] == False:
                    tAngLow, tAngHigh = -14, 90  # Eventually these should move to defaults page
                    try:
                        angInvalid = not isParamValid(
                            testAng, matchType="numeric", validValues=[tAngLow, tAngHigh]
                        )
                    except ValueError:  # hacky check for param type validation
                        angInvalid = True
                    if angInvalid:
                        invalidAcqParamText += (
                            f"\tInvalid Angle: {testAng} . For transmission samples, angle must be"
                            f" between {tAngLow} and {tAngHigh}\n"
                        )
                else:
                    angInvalid = True
                    raise ValueError("Unable to determine angle geometry, check grazing field")

                # if any invalid, need to report out
                invalidAcqParam = (
                    invalidAcqParam or angInvalid
                )  # we already had invalids, or we found one now

        if invalidAcqParam:
            warnings.resetwarnings()
            warnings.warn(invalidAcqParamText, stacklevel=2)

        # Pull data from PASS Database

        # Try to find proposal ID to
        if "proposal_id" in sam:
            sam["proposal_id"] = str(sam["proposal_id"])  # convert proposal id to a string
            proposal = sam["proposal_id"]
        elif "data_session" in sam:
            proposal = sam["data_session"]
        else:
            warnings.warn(
                "no valid proposal was located - please add that and try again", stacklevel=2
            )
            proposal = 0

        # Query the PASS database for values
        try:
            sam["data_session"], sam["analysis_dir"], sam["SAF"], sam["proposal"] = (
                get_proposal_info(proposal)
            )
        except:
            warnings.warn("PASS lookup failed - trusting values", stacklevel=2)
            pass

        if sam["SAF"] == None:
            print(f'line {i}, sample {sam["sample_name"]} - data will not be accessible')

        # Populate the "bar_loc" field
        new_bar[i]["bar_loc"]["spot"] = sam["bar_spot"]
        new_bar[i]["bar_loc"]["th"] = sam["angle"]

        # Get rid of the stupid unnamed columns thrown in by pandas
        for key in [key for key, value in sam.items() if "named" in key.lower() or "Index" in key]:
            del new_bar[i][key]

    if verbose:
        print("Bar and Acquisitions Sheets Loaded")
    return new_bar


def get_proposal_info(proposal_id, beamline="SST1", path_base="/sst/", cycle="2024-1"):
    """Query the api PASS database, and get the info corresponding to a proposal ID

    Parameters
    ----------
    proposal_id : str or int
        string of a number, a string including a "GU-", "PU-", "pass-", or  "C-" prefix and a number, or a number
    beamline : str, optional
        the beamline name from PASS, by default "SST1"
    path_base : str, optional
        the part of the path that indicates it's really for this beamline, by default "/sst/"
    cycle : str, optional
        the current cycle (or the cycle that is valid for this purpose), by default "2023-1"

    Returns
    -------
    tuple (res["data_session"], valid_path, valid_SAF, proposal_info)
         data_session ID which should be put into the run engine metadata of every scan, the path to write analyzed data to, the SAF, and all of the proposal information for the metadata if needed
    """

    warn_text = (
        "\n WARNING!!! no data taken with this proposal will be retrievable \n  it is HIGHLY"
        " suggested that you fix this \n if you are running this outside of the NSLS-II network,"
        " this is expected"
    )
    proposal_re = re.compile(r"^[GUCPpass]*-?(?P<proposal_number>\d+)$")
    if isinstance(proposal_id, str):
        proposal = proposal_re.match(proposal_id).group("proposal_number")
    else:
        proposal = proposal_id
    pass_client = httpx.Client(base_url="https://api-staging.nsls2.bnl.gov")
    responce = pass_client.get(f"/proposal/{proposal}")
    res = responce.json()
    if "safs" not in res:
        warnings.warn(
            f"proposal {proposal} does not appear to have any safs" + warn_text, stacklevel=2
        )
        pass_client.close()
        return None, None, None, None
    comissioning = 1
    if "cycles" in res:
        comissioning = 0
        if cycle not in res["cycles"]:
            warnings.warn(
                f"proposal {proposal} is not valid for the {cycle} cycle" + warn_text, stacklevel=2
            )
            pass_client.close()
            return None, None, None, None
    elif "Commissioning" not in res["type"]:
        warnings.warn(
            f"proposal {proposal} does not have a valid cycle, and does not appear to be a"
            " commissioning proposal"
            + warn_text,
            stacklevel=2,
        )
        pass_client.close()
        return -1
    if len(res["safs"]) < 0:
        warnings.warn(
            f"proposal {proposal} does not have a valid SAF in the system" + warn_text,
            stacklevel=2,
        )
        pass_client.close()
        return None, None, None, None
    valid_SAF = ""
    for saf in res["safs"]:
        if saf["status"] == "APPROVED" and beamline in saf["instruments"]:
            valid_SAF = saf["saf_id"]
    if len(valid_SAF) == 0:
        warnings.warn(
            f"proposal {proposal} does not have a SAF for {beamline} active in the system"
            + warn_text,
            stacklevel=2,
        )
        pass_client.close()
        return None, None, None, None
    proposal_info = res
    dir_responce = pass_client.get(f"/proposal/{proposal}/directories")
    dir_res = dir_responce.json()
    if len(dir_res) < 1:
        warnings.warn(f"proposal{proposal} have any directories" + warn_text, stacklevel=2)
        pass_client.close()
        return None, None, None, None
    valid_path = ""
    for dir in dir_res:
        if (path_base in dir["path"]) and (('commissioning' in dir["path"]) or (cycle in dir["path"])):
            valid_path = dir["path"]
    if len(valid_path) == 0:
        warnings.warn(
            f"no valid paths (containing {path_base} and {cycle} were found for proposal"
            f" {proposal}"
            + warn_text,
            stacklevel=2,
        )
        pass_client.close()
        return None, None, None, None

    pass_client.close()
    return res["data_session"], valid_path, valid_SAF, proposal_info


def save_samplesxlsx(bar, name="", path=""):
    """Exports the in-memory bar (list of sample dicts) as an excel sheet with 'Bar', and 'Acquisitions' sheets.
        exports with a fixed pattern to path out_date_name.xlsx
    Parameters
    ----------
    bar : list of dict
        list of sample dicts
    name : str
        export file name, e.g., test
    path : str/path
        export path
    """
    switch = {
        "RSoXS Sample Outboard-Inboard": "x",
        "RSoXS Sample Up-Down": "y",
        "RSoXS Sample Downstream-Upstream": "z",
        "RSoXS Sample Rotation": "th",
        "x": "x",
        "y": "y",
        "z": "z",
        "th": "th",
    }

    for samp in bar:
        for acq in samp.get("acq_history", []):
            for arg in acq["arguments"]:
                if isinstance(acq["arguments"][arg], np.ndarray):
                    acq["arguments"][arg] = list(acq["arguments"][arg])

    filename = path + f'out_{datetime.today().strftime("%Y-%m-%d_%H-%M-%S")}_{name}.xlsx'

    for samp in bar:
        for acq in samp.get("acq_history", []):
            for arg in acq["arguments"]:
                if isinstance(acq["arguments"][arg], np.ndarray):
                    acq["arguments"][arg] = list(acq["arguments"][arg])

    acqlist = []
    for i, sam in enumerate(bar):
        for acq in sam["acquisitions"]:
            acq.update({"sample_id": sam["sample_id"]})
            cleanacq = deepcopy(empty_acq)
            for key in acq:
                if isinstance(acq[key], (str)):
                    cleanacq[key] = acq[key]
                else:
                    cleanacq[key] = json.dumps(acq[key])
            acqlist.append(cleanacq)
    sampledf = pd.DataFrame.from_dict(bar, orient="columns")
    df_bar = deepcopy(sampledf)
    testdict = df_bar.to_dict(orient="records")
    cleanbar = []
    for i, sam in enumerate(testdict):
        if "acq_history" not in testdict[i].keys():
            testdict[i]["acq_history"] = []
        elif isinstance(testdict[i]["acq_history"], str):
            testdict[i]["acq_history"] = []
        # json dump the pythonic parts
        # including sample: bar_loc,location, proposal,acq_history
        testdict[i]["acq_history"] = json.dumps(testdict[i]["acq_history"])
        testdict[i]["bar_loc"] = json.dumps(testdict[i]["bar_loc"])
        testdict[i]["location"] = json.dumps(testdict[i]["location"])
        testdict[i]["proposal"] = json.dumps(testdict[i]["proposal"])
        del testdict[i]["acquisitions"]
        cleansam = deepcopy(empty_sample)
        cleansam.update(testdict[i])
        cleanbar.append(cleansam)

    sampledf = pd.DataFrame.from_dict(cleanbar, orient="columns")
    acqdf = pd.DataFrame.from_dict(acqlist, orient="columns")
    writer = pd.ExcelWriter(filename)
    sampledf.to_excel(writer, index=False, sheet_name="Bar")
    acqdf.to_excel(writer, index=False, sheet_name="Acquisitions")
    writer.close()

    excel_file = load_workbook(filename)
    excel_file.properties.title = current_version
    excel.save_workbook(excel_file, filename)
    excel_file.close()


def convertSampleSheetExcelMediaWiki(
    excelSheet: Path = None,
    paramsSheetToOutput: str = "all",
    rulesSheetName: str = "SheetRulesAndMetaData",
    versionCell: str = "B4",
    startRow_Params: int = 7,
    endRow_Params: int = None,
    startColumn_Params: str = "A",
    endColumn_Params: str = "F",
    verbose: bool = "TRUE",
) -> str:
    """Converts Sample Sheet Parameter Metadata into a MediaWiki-compatible format string.

    Parameters
    ----------
    excelSheet: Path
        Path (or string) to the excel sheet to be loaded.
    rulesSheetName: str
        Name of the excel sheet which should be parsed for metadata
    paramsSheetToOutput: str
        Which set of params should be output (e.g., 'bar', 'acquisitions'). 'all' will sequentially output tables for the same wiki page
    versionCell: str
        Location (e.g., 'B4') of the cell that contains the sheet version number
    startRow_Params: int
        Excel row number which contains the header for the metadata table (excel starts at row 1)
    endRow_Params: int
        Excel row number which contains the last row of metadata (leave as -1 if scanning to end of file)
    startColumn_Params: str
        First excel column (by letter) that contains the metadata table
    endColumn_Params:str
        Last excel column (by letter) that contains the metadata table
    verbose: bool
        Whether to print progress text to stdout


    Returns
    -------
    str
        A string containing the formatted table ready to copy-paste into MediaWiki
    """

    if verbose:
        print("-" * 5 + " Start Log" + "-" * 5)
        print("\tExtracting Sheet Version...", end=" ")

    # Step 1: Extract Version Cell and add to wiki table header
    ## Split versionCell in (row, column)
    versRow, versColumn = [
        int("".join(filter(str.isdigit, versionCell))),
        "".join(filter(str.isalpha, versionCell)),
    ]

    ## Extract Version Code as a string
    versionStr = load_workbook(excelSheet).properties.title
    # print(versionStr)

    ## Get Current Date
    date.today()

    ## Add Wiki Page Header to Output string
    outStr = (
        f"== SST-1 Sample Sheet Syntax Version: {versionStr} Last Updated: {date.today()} ==\n"
    )
    if verbose:
        print(f"Pass!\n\t\tVersion Number is -> {versionStr}")

    # Step 2: Extract Metadata Table from Excel Sheet
    if verbose:
        print("\tExtracting Sheet Metadata...")

    ## If endRow_Params is provided, limit the number of rows parsed
    if endRow_Params is None:
        numRows = None
    else:
        numRows = endRow_Params - startRow_Params

    ## Convert column bounds to string
    colString = startColumn_Params + ":" + endColumn_Params

    excelMetadataIn = pd.read_excel(
        excelSheet,
        sheet_name=rulesSheetName,
        header=startRow_Params - 1,
        nrows=numRows,
        usecols=colString,
    )

    ## Drop empty rows (where 'Sheet' is NaN)
    excelMetadataIn = excelMetadataIn.dropna(subset="Sheet")
    ## Replace NaNs and 'nan's with blank
    excelMetadataIn = excelMetadataIn.replace("nan", "")
    excelMetadataIn = excelMetadataIn.fillna(" ")

    ###display(excelMetadataIn)

    ## Build MediaWiki Tables

    ## Get list of unique sheets for which we have metadata
    sheetList = excelMetadataIn.Sheet.unique()

    if verbose:
        print(f"\t\tFound Metadata for these sheets: {sheetList}")
        print(f"\t\tUser requested tables for: {paramsSheetToOutput}")

    # Filter down the list of tables to output
    if paramsSheetToOutput.lower() == "all":
        sheetListToRun = sheetList
    else:
        sheetListToRun = [paramsSheetToOutput]

    if verbose:
        print(f"\t\tOutputting tables for: {sheetListToRun}...")

    # Create subset dataframes
    dataframeList = []
    for sheetName in sheetListToRun:
        dataframeList.append(excelMetadataIn[excelMetadataIn.Sheet == sheetName])

    ###print(dataframeList[1])

    ## Make one table per value in the 'Sheet' column
    for excelMetadataFrame in dataframeList:
        excelMetadataFrame.reset_index(drop=True, inplace=True)
        outStr += "\n" + r'{| class="wikitable sortable"' + "\n" + "|-\n"

        # Add header row elements
        outStr += "! "
        for colHeader in excelMetadataFrame.columns:
            filteredColHeader = str(colHeader).replace("\r", " ").replace("\n", " ")
            outStr += f"{filteredColHeader} !! "

        # trim extra "!! "
        outStr = outStr[:-4]

        ###display(excelMetadataFrame)

        # Add Metadata Row Elements

        ## Loop through metadata rows
        for mdRow in excelMetadataFrame.index:
            ###Loop through columns
            outStr += "\n|-\n| "
            ###print(mdRow)
            for mdVal in excelMetadataFrame.iloc[mdRow].to_list():
                filteredmdVal = str(mdVal).replace("\r", " ").replace("\n", " ")
                outStr += f"{filteredmdVal} || "
                ###print(f"\t{mdVal}")
            # trim extra " || " at the end of each line
            outStr = outStr[:-4]

        # Add MediaWiki Table End
        outStr += "\n|}\n"
        # print(outStr)

    if verbose:
        print("-" * 5 + " End Log. Copy text below this line into the wiki" + "-" * 5)

    return outStr


def isParamValid(
    testVal, matchType: str = "exact", validValues: list = None, invalidValues: list = None
):
    """Tests whether the testVal is valid based on matchType and both the validValues list and invalidValues list (if provided)

    Test against validValues is made first, then against invalidValues. You can use this to define complex numerical ranges.

    Parameters
    ----------
    testVal : any
        A single value to validate. Lists should be expanded outside of this function.
    matchType : str, optional
        Match method, by default 'exact'
        'exact' casts input as string returns True if input is in validValues and not in invalidValues
        'interval' attempts to cast input as float and returns True if value is within the validValues range and not in the invalidValues range
    validValues : list, optional
        A list of valid match values or a two-element list specifying a numeric interval to return True, by default None
    invalidValues : list, optional
        A list of invalid match values or a two-element list specifying a numeric interval to return False for, by default None
    """

    # Make sure parameter to validate is not a list
    # This rule is so that useful output is forced in the calling function
    if isinstance(testVal, list):
        raise ValueError("This function only validates single values, not lists")

    isValid = True

    # Check exact match
    if matchType.lower() == "exact":
        # Compare against validValues
        if validValues is not None:
            # cast validValues as str inside lists
            if not isinstance(validValues, list):
                validValues = [validValues]
            validValues = [str(i) for i in validValues]
            isValid = str(testVal) in validValues
        if invalidValues is not None:
            # cast validValues as str inside lists
            if not isinstance(invalidValues, list):
                invalidValues = [invalidValues]
            invalidValues = [str(i) for i in invalidValues]
            isValid = isValid and (str(testVal) not in invalidValues)

        return isValid
    elif matchType.lower() == "numeric":
        testVal = float(testVal)
        # Compare against validValues
        if validValues is not None:
            # Check that it is a two element list
            if not isinstance(validValues, list) or len(validValues) != 2:
                raise ValueError(
                    "If 'numeric' match is requested, validValues must be a 2-element list"
                    " specifying a numeric range"
                )
            # Cast validValues as float inside lists
            validValues = [float(i) for i in validValues]
            isValid = validValues[0] <= testVal and testVal <= validValues[1]
        # Compare against invalidValues
        if invalidValues is not None:
            # Check that it is a two element list
            if not isinstance(invalidValues, list) or len(invalidValues) != 2:
                raise ValueError(
                    "If 'numeric' match is requested, invalidValues must be a 2-element list"
                    " specifying a numeric range"
                )
            # Cast invalidValues as float inside lists
            invalidValues = [float(i) for i in invalidValues]
            isValid = isValid and not (invalidValues[0] <= testVal and testVal <= invalidValues[1])
    else:
        raise ValueError("Invalid matchType provided")
    return isValid
